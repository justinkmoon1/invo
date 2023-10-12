import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
import pandas as pd

file_path = r'C:\Users\Serry\Desktop\projects\py\invo\data\target price\target_price_sheet.xlsx'
wb = load_workbook(file_path)
ws = wb['Sheet1']

#utility function to download a webpage and return a beautiful soup doc
def get_page(url):
    response = requests.get(url)
    if not response.ok:
        print('Status code:', response.status_code)
        raise Exception('Failed to load page {}'.format(url))
    
    page_content = response.text
    doc = BeautifulSoup(page_content, 'html.parser')
    return doc

for i in range (2, 3):
    ticker = ws['A%d'%i].value

    my_url = "https://finance.yahoo.com/quote/"+ticker+"?p="+ticker
    response = requests.get(my_url)

    #function call
    doc = get_page(my_url)
    #appropritae tags common to news-headlines to filter out the necessary information.
    div_element = doc.find('div', class_='Pos(r) T(5px) Miw(100px) Fz(s) Fw(500) D(ib) C($primaryColor)Ta(c) Translate3d($half3dTranslate)')
    if div_element:
        # Find the <span> element within the <div>
        span_elements = div_element.find_all('span')
    
        # Check if the <span> element was found
    if len(span_elements) >= 2:
        # Access the second <span> element (index 1, since indexing starts from 0)
        second_span = span_elements[1]
    
        # Extract the text content from the second <span> element
        second_span_text = second_span.text
    
        # Print the text content of the second <span> element
    else:
        print("There are less than two <span> elements.")

    ws['B%d'%i].value = second_span_text
    wb.save(file_path)

#Catching Exceptions
#print("response.ok : {} , response.status_code : {}".format(response.ok , response.status_code))
#print("Preview of response.text : ", response.text[:5000])