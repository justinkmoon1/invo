import openpyxl
from openpyxl import workbook, load_workbook
from datetime import date
today = date.today()

import yfinance as yf
import pandas as pd

# Use a raw string (prefix 'r') or escape backslashes in the path
file_path = r'C:\Users\Serry\Desktop\projects\py\invo\data\excell automation\totalValuation\Total Valuation Model.xlsx.xlsx'

wb = load_workbook(file_path)
ws = wb['Master Inputs Start here']

#set company name + TICKER!!!!
ws['B4'] = input("enter name of company: ")
ticker = yf.Ticker(input("name in yfinance: "))
#set date of valuation
ws['B5'] = today.strftime("%m/%d/%y")
print(ws['B5'])
#set country of incorporation
ws['B9'] = ticker.info['country']


#Current Operating Income
ws['B12'] = 
#Current Interest Expense 
ws['B13'] =
#Current Capital Spending
ws['B14'] = 
#Current Depreciation & Amort'n 
ws['B15'] =
#Effective tax rate (for use on operating income)
ws['B16'] = 
#Marginal tax rate (for use on cost of debt)
ws['B17'] = 
#Current Revenues + previous year
ws['B18'] = 
ws['C18'] = 
#Current Non-cash Working Capital
ws['B19'] = 
#Chg. Working Capital 
ws['B20'] = 
#Book Value of Debt + previous year
ws['B21'] = 
ws['C21'] = 
#Book Value of Equity + previous year
ws['B22'] = 
ws['C22'] = 
#Cash & Marketable Securities + previous year
ws['B24'] = 
ws['C24'] = 
#Value of Non-operating Assets + previous year
ws['B25'] = 
ws['C23'] = 
#Minority interests + previous year
ws['B26'] = 
ws['C26'] = 

#currently traded
a = input("Is your stock currently traded? (Y/N): ")
if (a=="Y"):
    ws['B29'] = "Yes"
    ws['B31'] = 
    ws['B32'] = 
    ws['B33'] = 
    ws['B35'] = ""
    ws['B36'] = ""


#General Market Data	
#Long Term Riskfree rate
ws['B39'] = ""
#Equity Risk Premium
ws['B40'] = ""

#Ratings	
#Do you want to estimate the firm's synthetic rating =	Yes
ws['B43'] = ""
#If yes, choose the type of firm
ws['B44'] = 
#If not, what is the current rating of the firm?
ws['B45'] = 
#Enter the cost of debt associated with the rating 
ws['B46'] = 
	
#Options	
#Do you have equity options (management options, warrants) outstanding?	No
ws['B49'] = 
#If yes, enter the number of options
ws['B50'] = 
#Average strike price 
ws['B51'] = 
#Average maturity
ws['B52'] = 
#Standard Deviation in stock price
ws['B53'] = 
#Do you want to use the stock price to value the 
#option or your estimated value?
ws['B54'] = 

	
#Valuation Inputs	
#High Growth Period	The questions below, especially the yes or no ones, can be confusing. Please read the comments on the input cells.
#Length of high growth period =	10
ws['B58'] = 
#Beta to use for high growth period for your firm=	1.14
ws['B59'] = 
67
#Do you want to keep the debt ratio computed from your inputs?	Yes
ws['B60'] = 
#If yes, the debt ratio that will be used to compute the cost of capital is	19.21%
ws['B61'] = 
#If no, enter the debt ratio that you would like to use in the high growth period	7.00%
ws['B62'] = 
#Do you want to keep the existing ratio of working capital to revenue?	Yes
ws['B63'] = 
#If yes, the working capital as a percent of revenues will be	6.90%
ws['B64'] = 
#If no, enter the ratio of working capital to revenues to use in analysis	12%
ws['B65'] = 
#Do you want to compute your growth rate from fundamentals?	Yes
ws['B66'] = 
#If no, enter the expected growth rate in operating income for high growth period	15%
ws['B67'] = 
#If yes, the inputs to the fundamental growth calculation (based upon your inputs) are	
#Do you want to change these inputs?	No
ws['B71'] = 
#Return on Capital =	33.00%
ws['B72'] = 
#Reinvestment Rate =	51.11%
ws['B73'] = 

#Do you want me to gradually adjust your high growth inputs in the second half?	Yes
ws['B75'] = 

#Stable Growth Period	
#Growth rate during stable growth period =	3.00%
ws['B78'] = 
#Beta to use in stable growth period =	1.00
ws['B79'] = 
#Equity Risk Premium in stable growth =	7.50%
ws['B80'] = 
#Debt Ratio to use in stable growth period =	30.00%
ws['B81'] = 
#Pre-tax cost of debt in stable growth period =	9.00%
ws['B82'] = 
#Tax Rate to use in stable growth period =	34.00%
ws['B83'] = 
#To compute the reinvestment rate in stable growth, you have two options	
#Do you want to compute reinvestment needs in stable growth based on fundamentals?	Yes
ws['B85'] = 
#If yes, enter the return on capital that the firm will have in stable growth	25.00%
ws['B86'] = 
#If no,  enter capital expenditure as % of depreciation in stable growth	120%
ws['B87'] = 



